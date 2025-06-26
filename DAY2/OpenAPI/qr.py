import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re


filePath = r'고등교육기관 하반기 주소록(2024).xlsx'
df_from_excel = pd.read_excel(filePath,engine='openpyxl')
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0,5)))


url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'   #도로명주소
road_type2 = 'PARCEL' #지번주소
address = '&address='
keys = '&key='
primary_key = '45387057-B873-3138-8FC4-0DBB9084DA60'

def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x,y
    else:
        x = 0
        y = 0
        return x,y


try:
    wb = load_workbook(r"학교주소좌표.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

for num,value in enumerate(address_list):
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    x,y = request_geo(addr)
    sheet.append([university_list[num],addr,x,y])

wb.save(r"학교주소좌표.xlsx")

import folium

map = folium.Map(location=[37,127],zoom_start=7)

marker = folium.Marker([35.183778757,126.880647243],
                       popup='한국폴리텍대학 광주캠퍼스',
                       icon = folium.Icon(color='blue'))

marker.add_to(map)

map.save(r'poly_map.html')